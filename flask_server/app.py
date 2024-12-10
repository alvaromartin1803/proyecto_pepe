from flask import Flask, request, render_template
import openpyxl

app = Flask(__name__)

def credenciales_duplicadas(usuario, contraseña, hoja):
    for row in hoja.iter_rows(min_row=2, values_only=True):
        registrado_usuario, registrado_contraseña = row
        if registrado_usuario == usuario and registrado_contraseña == contraseña:
            return True
    return False

def ajustar_anchura_columnas(hoja):
    for column_cells in hoja.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        hoja.column_dimensions[column_cells[0].column_letter].width = length + 2

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    usuario = request.form['username']
    contraseña = request.form['password']
    
    try:
        workbook = openpyxl.load_workbook('registro.xlsx')
        hoja = workbook.active
    except openpyxl.utils.exceptions.InvalidFileException:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = 'Credenciales'
        hoja.append(['Usuario', 'Contraseña'])
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = 'Credenciales'
        hoja.append(['Usuario', 'Contraseña'])

    if credenciales_duplicadas(usuario, contraseña, hoja):
        return 'Credenciales ya registradas'

    hoja.append([usuario, contraseña])
    ajustar_anchura_columnas(hoja)  
    workbook.save('registro.xlsx')
    
    return 'Registro exitoso'

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
